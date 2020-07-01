"""
用于读取Excel文件，数据在 H:\DATA\data1.xlsx 文件的sheet3表格中
"""
# 处理表格模块
import openpyxl
# 目录控制模块
import os
# 日志模块
import logging
# 漂亮打印
import pprint

'''
日志设置
'''
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')

'''
全局变量
'''
# 工作目录(data文件夹)
WORKPATH = 'H:/DATA'
# 项目命名
SHORTNAME1 = ['U1', 'U2', 'U3', 'U4', 'U5', 'U6', 'U7', 'U8', 'U9', 'U10', 'U11', 'U12', 'U13', 'U14', 'U15', 'U16',
              'U17', 'U18', 'U19']
SHORTNAME2 = ['D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D9', 'D10', 'D11', 'D12', 'D13', 'D14', 'D15', 'D16',
              'D17', 'D18', 'D19']
# 命名转换
NAME = {'U1': '房价相对上升', 'U2': '旅游花费相对上升', 'U3': '铁路运输占比上升',
        'U4': '民航运输占比上升', 'U5': '国企经济占比上升', 'U6': '私企经济占比上升',
        'U7': '出生率上升', 'U8': '死亡率上升', 'U9': '第一产业占比上升',
        'U10': '第二产业占比上升', 'U11': '第三产业占比上升', 'U12': '进口占比上升',
        'U13': '出口占比上升', 'U14': '对美汇率上升', 'U15': '财政收入增速上升',
        'U16': '财政支出增速上升',
        'U17': '人均收入增速上升', 'U18': '旅游花费增速上升', 'U19': '房价增速上升',

        'D1': '房价相对下降', 'D2': '旅游花费相对下降', 'D3': '铁路运输占比下降',
        'D4': '民航运输占比下降', 'D5': '国企经济占比下降', 'D6': '私企经济占比下降',
        'D7': '出生率下降', 'D8': '死亡率下降', 'D9': '第一产业占比下降',
        'D10': '第二产业占比下降', 'D11': '第三产业占比下降', 'D12': '进口占比下降',
        'D13': '出口占比下降', 'D14': '对美汇率下降', 'D15': '财政收入增速下降',
        'D16': '财政支出增速下降',
        'D17': '人均收入增速下降', 'D18': '旅游花费增速下降', 'D19': '房价增速下降'
        }
# 存储结构
dataStruct = []

'''
工作目录设置
'''
# 改变工作目录
os.chdir(WORKPATH)


def loadData():
    '''
    读取excel
    '''
    # 打开文件
    workbook = openpyxl.load_workbook('data1.xlsx')
    # 获取Sheet2表格
    sheet = workbook['Sheet3.2']
    '''
    get_sheet_by_name()方法已弃用
    sheet = workbook.get_sheet_by_name('Sheet2')
    '''
    # 获取数据,按行读取，两层循环，外层循环行，内层循环列
    for i in range(2, sheet.max_row + 1):
        temp = []
        for j in range(2, sheet.max_column + 1):
            data = sheet.cell(i, j).value
            # 1升0降
            if data == 1:
                # 下面两行第一行使用代号计算，第二行是用汉字计算，结果相同
                # temp.append(SHORTNAME1[j - 2])
                temp.append(NAME[SHORTNAME1[j - 2]])
            else:
                # temp.append(SHORTNAME2[j - 2])
                temp.append(NAME[SHORTNAME2[j - 2]])
        dataStruct.append(temp)
    # pprint.pprint(dataStruct)
    # print(len(dataStruct))
    return dataStruct


# loadData()
