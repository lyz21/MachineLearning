# encoding=utf-8
"""
Name:         analysisemotion
Description:  
Author:       LiuYanZhe
Date:         2019/11/8
"""

import openpyxl
# 操作目录
import os
# 情感分析
from snownlp import SnowNLP
# 分词器
import jieba

'''全局变量'''
# 工作目录
WORKDIR = 'E:/data_lyz'
# 文件名
FILENAME = '我和我的祖国情感分析.xlsx'
# 设置工作目录
os.chdir(WORKDIR)


# 分析所有评论情感
def all_emotion():
    # 获取文件
    workbook = openpyxl.load_workbook(FILENAME)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")
    # 存储情感
    list=[]
    for i in range(2,sheet.max_row+1):
        comm=sheet.cell(i,2).value
        emotion_score=SnowNLP(comm).sentiments
        print(emotion_score)
# 分析包含“祖国”的评论情感
def nation_emotion():
    # 获取文件
    workbook = openpyxl.load_workbook(FILENAME)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")
    # 存储情感
    list=[]
    for i in range(2,sheet.max_row+1):
        # 获取评论
        comm=sheet.cell(i,2).value
        # 分词
        short_list=jieba.cut(comm, cut_all=True)
        # 包含中国
        if '祖国' in short_list or '中国' in short_list or '国家' in short_list:
            emotion_score=SnowNLP(comm).sentiments
            # print(comm,':',emotion_score)
            print(emotion_score)
# 分析包含“陈凯歌”的评论情感
def kaige_emotion():
    # 获取文件
    workbook = openpyxl.load_workbook(FILENAME)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")
    # 存储情感
    list=[]
    for i in range(2,sheet.max_row+1):
        # 获取评论
        comm=sheet.cell(i,2).value
        # print(comm)
        # 分词
        short_list=jieba.cut(comm, cut_all=True)
        # 包含中国
        # if '凯歌' in short_list or '陈凯歌' in short_list:
        # if '峥' in short_list or '徐峥' in short_list:
        if ('宁' in short_list) or ('峥' in short_list) or ('渤' in short_list) or ('葛' in short_list):
            emotion_score=SnowNLP(comm).sentiments
            # print(comm,':',emotion_score)
            print(emotion_score)

# nation_emotion()
kaige_emotion()