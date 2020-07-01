# encoding=utf-8
"""
Name:         analyTime
Description:  分析时光网影评
Author:       LiuYanZhe
Date:         2019/11/2
"""
# 操作Excel表格
import openpyxl
# 操作目录
import os
# 日志
import logging
# 此处用作保存数据
import numpy
# 拆分
import jieba
# 导入Apriori算法（自己写的）
import DataAnalysisMovie.Apriori as apriori

'''日志设置'''
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# 禁用日志
# logging.disable()

'''全局变量'''
# 工作目录
WORKDIR = 'E:/data_lyz'
# 文件名
FILENAME = '我和我的祖国时光网影评T.xlsx'
# 保存的文件名
FILENAME_NEW = FILENAME.replace('.xlsx', '')
# 无意义词
NOMEANING_LIST = ['', '也', '了', '的', '你', '在', '都', '，', '？', '是', '故事', '一个', '有', '我', '这', '流星']
# 关键词
KEYWORDS_LIST = []
# 设置工作目录
os.chdir(WORKDIR)

# 加载关键词
def loadKeyWors():
    # 获取文件
    workbook = openpyxl.load_workbook('keyWordslist.xlsx')
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")
    keyWords_list=[]
    for i in range(1, sheet.max_row + 1):
        keyWord=sheet.cell(i,1).value
        keyWords_list.append(keyWord)
    return keyWords_list
# 评论类
# class Comment:
#     def __init__(self, userName,score, content,date):
#         self.score = score  # 评分
#         self.date = date  # 评论日期
#         self.content = content  # 评论内容
# 加载数据
# 返回评论类列表，存储结构data_list[temp_list[]]
def loadData(filename):
    # 获取文件
    workbook = openpyxl.load_workbook(filename)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")
    # 数据保存到list
    data_list = []
    for i in range(2, sheet.max_row + 1):
        temp_list = []
        userName = sheet.cell(i, 1).value
        temp_list.append(userName)
        score = sheet.cell(i, 2).value
        temp_list.append(score)
        comment = sheet.cell(i, 3).value
        temp_list.append(comment)
        date = sheet.cell(i, 4).value
        temp_list.append(date)
        data_list.append(temp_list)
    return data_list


# 拆分传入的集合的评论，并将词存入一个集合返回
def splitComment(list_data):
    split_list = []
    for i in range(len(list_data)):
        comment = list_data[i][2]
        temp_list = jieba.cut(comment, cut_all=False)
        for comm in temp_list:
            split_list.append(comm)
    return split_list


# 拆分传入的集合的评论，并将词存入两个集合返回
def splitComment2(list_data):
    global KEYWORDS_LIST
    split_list = []
    for i in range(len(list_data)):
        comment = list_data[i][2]
        score=list_data[i][1]
        temp0_list = []
        temp_list = jieba.cut(comment, cut_all=True)
        for comm in temp_list:
            if comm=='葛':
                comm='葛优'
            if comm=='凯歌' or comm=='白昼':
                comm='陈凯歌'
            if comm=='峥' or comm=='夺冠':
                comm='徐峥'
            # if comm not in KEYWORDS_LIST:
            #     continue
            temp0_list.append(comm)
        temp0_list.append(str(score))
        split_list.append(temp0_list)
    return split_list


# 统计分词数目
def count_split(split_list):
    split_count_dic = {}
    for split in split_list:
        split_count_dic.setdefault(split, 0)
        split_count_dic[split] += 1
    return split_count_dic


# 筛选评分段对应的集合
def getscoreX(data_list, score_min, score_max):
    result_list = []
    for entry in data_list:
        if float(entry[1]) >= score_min and float(entry[1]) < score_max:
            result_list.append(entry)
    return result_list


def main():
    global KEYWORDS_LIST
    # 加载数据
    data_list = loadData(FILENAME)
    # 加载关键词
    KEYWORDS_LIST=loadKeyWors()
    '''     '''
    # # 筛选评10-50分的评论
    # listByScore=getscoreX(data_list,5,8)
    # # 将评论分词
    # split_list=splitComment(listByScore)
    # # 统计分词数目
    # split_count_dic=count_split(split_list)
    # for k,v in split_count_dic.items():
    #     if v<=1:
    #         continue
    #     print(k,':',v)

    ''''# Apriori    '''
    # 将评论分词，返回二维数组
    split_list2 = splitComment2(data_list)
    print(split_list2)
    # 使用Apriori算法分析
    apriori.Apriori(split_list2)

    # list=numpy.load(FILENAME+'.npy')


main()
