# encoding=utf-8
"""
Name:         countNumByScore
Description:  计算各分数段评分数目
Author:       LiuYanZhe
Date:         2019/11/6
"""
import openpyxl
# 操作目录
import os

'''全局变量'''
# 工作目录
WORKDIR = 'E:/data_lyz'
# 文件名
FILENAME = '分析评分.xlsx'
# 设置工作目录
os.chdir(WORKDIR)


# 统计各分数评分人数比例
def countRateByScore():
    # 获取文件
    workbook = openpyxl.load_workbook(FILENAME)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")
    count_score = {}
    for i in range(2, sheet.max_row + 1):
        score = sheet.cell(i, 1).value
        if score=='comment-time':
            continue
        count_score.setdefault(score, 0)
        count_score[score] += 1
    for k, v in count_score.items():
        print(str(k), ' ', str(v))


# 统计高赞评论各分数评分人数比例
def countRateByHighScore():
    # 获取文件
    workbook = openpyxl.load_workbook(FILENAME)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet = workbook.get_sheet_by_name("Sheet1")

    count_score = {}
    for i in range(2, sheet.max_row + 1):
        # 赞数
        good = sheet.cell(i, 2).value
        if good==None:
            break
        if int(good) < 100:
            continue
        score = sheet.cell(i, 1).value
        if score=='comment-time':
            continue
        count_score.setdefault(score, 0)
        count_score[score] += 1
    for k, v in count_score.items():
        print(str(k), ' ', str(v))

# countRateByScore()
countRateByHighScore()
