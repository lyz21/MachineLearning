# encoding=utf-8
"""
Name:         analyNature
Description:  分析豆瓣评论（包括用户信息）
Author:       LiuYanZhe
Date:         2019/11/1
"""
# 操作Excel表格
import openpyxl
# 操作目录
import os
# 日志
import logging
# 此处用作保存数据
import numpy

'''日志设置'''
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# 禁用日志
# logging.disable()

'''全局变量'''
# 工作目录
WORKDIR = 'E:/data_lyz'
# 文件名
# FILENAME = '我和我的祖国豆瓣影评T.xlsx'
FILENAME = '红海行动豆瓣影评T.xlsx'
# FILENAME = '三块广告牌豆瓣影评T.xlsx'
# FILENAME = '至暗时刻豆瓣影评T.xlsx'
# 表格名
SHEETNAME = '我和我的祖国豆瓣'
# SHEETNAME = '红海行动豆瓣'
# SHEETNAME = '三块广告牌豆瓣'
# SHEETNAME = '至暗时刻豆瓣'
# 美国阵营
ENLIST = ['美国', '日本', '韩国', '德国', '西德', '法国', '澳大利亚', '英国', '瑞士', '意大利', '瑞典', '新加坡', '新西兰']

# 设置工作目录
os.chdir(WORKDIR)


# 用户类
class User:
    def __init__(self, ID, area):
        self.ID = ID  # 用户ID
        self.area = area  # 用户地区
        self.likeMovie = []  # 喜欢的电影列表
        self.likeMovieArea = []  # 喜欢的电影地区

    # 添加喜欢的电影方法
    def addMovie(self, likeMovie):
        self.likeMovie.append(likeMovie)

    # 添加喜欢的电影地区方法
    def addMovieArea(self, likeMovieArea):
        self.likeMovieArea.append(likeMovieArea)

    # toString方法
    def toString(self):
        string = '用户id；' + str(self.ID) + '； 用户地区；' + str(self.area) + '； 用户喜欢的电影；' + str(
            self.likeMovie) + '； 用户喜欢的电影所属地区；' + str(self.likeMovieArea)
        return string


# 评论类
class Comment:
    def __init__(self, score, date, content, good, user):
        self.score = score  # 评分
        self.date = date  # 评论日期
        self.content = content  # 评论内容
        self.good = good  # 点赞人数
        self.user = user  # 用户信息

    # toString方法
    def toString(self):
        string = '评分；' + str(self.score) + '； 评论日期：' + str(self.date) + '； 评论内容：' + str(self.content) + '； 点赞人数；' + str(
            self.good) + self.user.toString()
        return string


# 去除字符串空格、换行方法
def delspace(string):
    # 去除换行
    string = string.replace('\n', '')
    # 去除空格
    string = string.replace(' ', '')
    string = string.replace('\xa0', '')
    # 去除两边空格
    string = string.strip()
    return string


# 统计user喜欢的电影各地区数，存储结构dic_area[地区：统计值]
def get_likeAreaCount(user):
    likeMovieArea_list = user.likeMovieArea
    dic_area = {}
    china_count = 0
    en_count = 0  # 美国阵营所占比例
    am_count = 0  # 美国电影占比
    all_count = 0
    for areas in likeMovieArea_list:
        area_list = areas.split('/')
        for i in range(len(area_list)):
            area = area_list[i]
            if area == '中国大陆' or area == '中国香港' or area == '中国台湾' or area == '台湾' or area == '香港' or area == '中国':
                area = '中国'
                china_count += 1
            if area in ENLIST:
                en_count += 1
            if area == '美国':
                am_count += 1
            dic_area.setdefault(area, 0)
            dic_area[area] += 1
            all_count += 1
        # 中国电影所占比例
        china_ratio = china_count / 10
        # 美国阵营电影比例
        en_ratio = en_count / all_count
        # 美国电影占比
        am_ratio = am_count / 10
    return dic_area, china_ratio, en_ratio, am_ratio, china_count


# 统计所有评论的用户各地区人数
def get_userPosition(list_comments):
    dic_pos = {}
    for comment in list_comments:
        dic_pos.setdefault(comment.user.area, 0)
        dic_pos[comment.user.area] += 1
    return dic_pos
# 统计每个分数段的各地区用户人数,dic_info{评分：list[评论对象]},返回值数据结构dic_1{分数：dic{地区:统计值}}
def get_userPositionByearea(dic_info):
    # 创建workbook对象
    workbook = openpyxl.Workbook()
    # 获取当前sheet对象（表格）
    sheet = workbook.get_active_sheet()
    sheet['A1']='分数'
    sheet['B1']='地区'
    sheet['C1']='数量'
    row=2
    for i in range(10,60,10):
        v=dic_info[str(i)]
        dic_2={}
        for j in range(len(v)):
            dic_2.setdefault(v[j].user.area,0)
            dic_2[v[j].user.area]+=1
        for k,va in dic_2.items():
            sheet['A' + str(row)] = str(i)
            sheet['B' + str(row)] = k
            sheet['C' + str(row)] = va
            row+=1
    # 设置表名,标明为页号
    sheet.title = str(SHEETNAME)
    # 保存文件
    workbook.save('按评分分类地区-' + FILENAME)
'''读取文件'''


# 返回评论类列表，存储结构list_comments[评论对象]
def loadData(filename):
    # 获取文件
    workbook = openpyxl.load_workbook(filename)
    # 获取当前表格
    # sheet = workbook.get_active_sheet()
    sheet=workbook.get_sheet_by_name("Sheet1")
    # 读取数据
    list_comments = []
    # 外层循环，循环行,将数据存储在类中
    for i in range(2, sheet.max_row + 1):
        # cell（行，列）
        score = sheet.cell(i, 1).value
        date = sheet.cell(i, 2).value
        content = sheet.cell(i, 3).value
        content = delspace(content)
        good = sheet.cell(i, 4).value
        user_Area = sheet.cell(i, 5).value
        user_ID = sheet.cell(i, 6).value
        user = User(user_ID, user_Area)
        for j in range(7, 17):
            like_movie = sheet.cell(i, j).value
            if like_movie == None:
                like_movie = '无'
            user.likeMovie.append(like_movie)
        for j in range(17, 27):
            like_movie_area = sheet.cell(i, j).value
            if like_movie_area == None:
                like_movie_area = '无'
            user.likeMovieArea.append(like_movie_area)
        comment = Comment(score, date, content, good, user)
        # logging.debug(comment.toString())
        list_comments.append(comment)
    return list_comments


'''找出各个评分用户信息'''


# 加载评论类列表，返回值存储结构：dic_num[评分:统计数]，dic_info[评分：list[评论对象]]
def getUserByScore(list_comments):
    # 存储每个评分人数
    dic_num = {}
    # 存储每个评分的评论类
    dic_info = {}
    for i in range(len(list_comments)):
        score = list_comments[i].score
        dic_num.setdefault(score, 0)
        dic_num[score] += 1
        templist = []
        dic_info.setdefault(score, templist)
        dic_info[score].append(list_comments[i])
    return dic_num, dic_info


'''按评分分析用户喜爱电影地区'''


# 参数存储结构：dic_info[评分：list[评论对象]]
def analyByScore(dic_info):
    # 创建workbook对象
    workbook = openpyxl.Workbook()
    # 获取当前sheet对象（表格）
    sheet = workbook.get_active_sheet()
    row_num = 2  # 记录行数
    for i in range(10, 60, 10):
        sheet['A1'] = '评分'
        sheet['B1'] = '中国电影平均占比'
        sheet['C1'] = '美国阵营电影平均占比'
        sheet['D1'] = '美国电影平均占比'
        sheet['E1'] = '中国电影平均个数'
        comments_list = dic_info[str(i)]
        logging.debug('评分为' + str(i))
        sheet['A' + str(row_num)] = str(i)
        china_avg_ratio = 0  # 计算中国电影平均占比
        en_avg_ratio = 0  # 计算美国阵营电影平均占比
        am_avg_ratio = 0  # 计算美国电影平均占比
        china_avg_count = 0  # 计算中国电影平均个数
        all_count = 0
        for comment in comments_list:
            if comment.user.likeMovie[0] == '无':
                continue
            dic_area, china_ratio, en_ratio, am_ratio, china_count = get_likeAreaCount(comment.user)
            china_avg_ratio = china_avg_ratio + china_ratio
            en_avg_ratio = en_ratio + en_avg_ratio
            am_avg_ratio = am_avg_ratio + am_ratio
            china_avg_count = china_avg_count + china_count
            all_count += 1
            # logging.debug('喜欢的电影所在地区：' + str(dic_area))
            # logging.debug('中国电影所占比例：'+str(china_ratio))
            # logging.debug('美国阵营电影所占比例：'+str(en_ratio))

        china_avg_ratio = china_avg_ratio / all_count
        en_avg_ratio = en_avg_ratio / all_count
        am_avg_ratio = am_avg_ratio / all_count
        china_avg_count = china_avg_count / all_count
        logging.debug('中国电影平均所占比例：' + str(china_avg_ratio))
        sheet['B' + str(row_num)] = str(china_avg_ratio)
        logging.debug('美国阵营电影平均所占比例：' + str(en_avg_ratio))
        sheet['C' + str(row_num)] = str(en_avg_ratio)
        # logging.debug('美国电影平均所占比例：' + str(am_avg_ratio))
        sheet['D' + str(row_num)] = str(am_avg_ratio)
        logging.debug('中国电影平均个数：' + str(china_avg_count))
        sheet['E' + str(row_num)] = str(china_avg_count)
        row_num += 1
    # 设置表名,标明为页号
    sheet.title = str('按评分分类分析')
    # 保存文件
    workbook.save('按评分分类分析-' + FILENAME)


'''写入表格方法'''


def writeExcel(data, filename,flag):
    # 创建workbook对象
    workbook = openpyxl.Workbook()
    # 获取当前sheet对象（表格）
    sheet = workbook.get_active_sheet()
    # 写入字典
    if flag==1:
        row=1
        for k, v in data.items():
            sheet['A'+str(row)]=k
            sheet['B'+str(row)]=v
            row+=1
    # 设置表名,标明为页号
    sheet.title = str(FILENAME)
    # 保存文件
    workbook.save(filename+'-'+ FILENAME)

'''主方法'''


def main():
    # 加载数据
    list_comments = loadData(FILENAME)
    '''按评分分类分析'''
    # 找出各个评分人数及评论信息
    idc_num, dic_info = getUserByScore(list_comments)
    #按评分分析用户喜爱电影地区
    # analyByScore(dic_info)
    # 汇总评论的各个地区人数
    # dic_pos = get_userPosition(list_comments)
    # writeExcel(dic_pos,'总的地区分布',1)
    # 找出各个分数的各个地方人数
    get_userPositionByearea(dic_info)

main()
