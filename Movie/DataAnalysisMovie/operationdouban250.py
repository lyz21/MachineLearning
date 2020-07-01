# 处理表格
import openpyxl
# 目录控制
import os
# 日志
import logging

'''日志设置'''
logging.basicConfig(level=logging.DEBUG,format=' %(asctime)s - %(levelname)s - %(message)s')

'''
全局变量
'''
# 工作目录
WORKDIR='H:/python/bookTask'
# 打开的文件名
FILENAME='豆瓣电影250_2.xlsx'
# 保存的文件名
FILENAMENEW='豆瓣电影250_2_operate.xlsx'
# 表名
SHEETTITLE=''
list_column=['A','B','C','D','E','F','G','H','I']
# 存储出现的地区
LIST_PLACE=[]
DIC_PLACE={}
# 存储出现的类型
LIST_TYPE=[]
DIC_TYPE={}

'''数据结构''' 
# 存储电影名，作为索引
list_movie_name=[]
# 存储{电影名:dic_movie_detail}
dic_movie_name={}
# 电影信息{'导演':'',...}
dic_movie_detail={}

# 改变工作目录
os.chdir(WORKDIR)

'''读取excel'''
# 打开文件
workbook=openpyxl.load_workbook(FILENAME)
# 获得当前表格
worksheet=workbook.get_active_sheet()
# 获得表名
SHEETTITLE=worksheet.title
'''
# 读取数据，存入数据结构中
for i in range(2,worksheet.max_row+1):
	for j in range(1,10):
		if j=1:
			# 将电影名加入list
			movie_name=worksheet.cell(i,j)
			list_movie_name.append(movie_name)
'''
# 读取地区、类型一列，存入列表
for i in range(2,worksheet.max_row+1):
	# 地区
	place=worksheet.cell(i,7).value
	place_list=place.split('/')
	LIST_PLACE=LIST_PLACE+place_list
	# 类型
	type=worksheet.cell(i,8).value
	type_list=type.split('/')
	LIST_TYPE=LIST_TYPE+type_list
	
# 统计地区数目
for i in range(len(LIST_PLACE)):
	# 去空格
	LIST_PLACE[i]=LIST_PLACE[i].replace(' ','')
	LIST_PLACE[i]=LIST_PLACE[i].strip()
	# 多义词合并
	if '中国大陆' in LIST_PLACE[i]:
		LIST_PLACE[i]='中国'
	elif '香港' in LIST_PLACE[i]:
		LIST_PLACE[i]='中国香港'
	elif LIST_PLACE[i]==' ' or len(LIST_PLACE[i])==0:
		continue
	# 地区存储到字典并统计数目
	DIC_PLACE.setdefault(LIST_PLACE[i],0)
	DIC_PLACE[LIST_PLACE[i]]=DIC_PLACE[LIST_PLACE[i]]+1

# 统计类型数目
for i in range(len(LIST_TYPE)):
	# 去空格
	LIST_TYPE[i]=LIST_TYPE[i].replace(' ','')
	LIST_TYPE[i]=LIST_TYPE[i].strip()
	if LIST_TYPE[i]==' ' or len(LIST_TYPE[i])==0:
		continue
	# 类型存储到字典并统计数目
	DIC_TYPE.setdefault(LIST_TYPE[i],0)
	DIC_TYPE[LIST_TYPE[i]]=DIC_TYPE[LIST_TYPE[i]]+1
logging.debug(DIC_PLACE)
'''写入Excle'''
# 创建workbook对象
workbook2=openpyxl.Workbook()
# 获取当前表
worksheet2=workbook2.get_active_sheet()
# 行号标记
j=2
# 地区循环赋值
for k,v in DIC_PLACE.items():
	worksheet2['A'+str(j)]=k
	worksheet2['B'+str(j)]=v
	j+=1
# 重置行号
j=2
# 类型循环赋值
for k,v in DIC_TYPE.items():
	worksheet2['D'+str(j)]=k
	worksheet2['E'+str(j)]=v
	j+=1
# 设置表名，与原表一致
worksheet2.title=SHEETTITLE
# 保存文件
workbook2.save(FILENAMENEW)
print('完成！')











